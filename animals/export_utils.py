from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime


def export_animals_to_excel(queryset, filename=None):
    """Export animals queryset to Excel file"""
    
    if filename is None:
        filename = f'animals_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ζώα"
    
    # Define headers
    headers = [
        'Chip ID', 'Όνομα', 'Είδος', 'Φύλο', 'Ηλικία', 
        'Συμπεριφορά', 'Στείρωση', 'Τραυματισμένο', 'Κλουβί',
        'Τοποθεσία Εύρεσης', 'Ημ/νία Εύρεσης', 'Ημ/νία Εισόδου',
        'Κατάσταση Υιοθεσίας', 'Δημόσια Προβολή', 'Καταφύγιο',
        'Δημιουργήθηκε Από', 'Δημιουργήθηκε Στις'
    ]
    
    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_num, animal in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1).value = animal.chip_id
        ws.cell(row=row_num, column=2).value = animal.name
        ws.cell(row=row_num, column=3).value = animal.get_species_display()
        ws.cell(row=row_num, column=4).value = animal.get_gender_display()
        
        # Age
        if animal.age_numeric:
            ws.cell(row=row_num, column=5).value = f"{animal.age_numeric} έτη"
        elif animal.age_category:
            ws.cell(row=row_num, column=5).value = animal.get_age_category_display()
        else:
            ws.cell(row=row_num, column=5).value = "Μη καθορισμένη"
        
        ws.cell(row=row_num, column=6).value = animal.get_behavior_display() if animal.behavior else ""
        ws.cell(row=row_num, column=7).value = animal.get_sterilization_status_display() if animal.sterilization_status else ""
        ws.cell(row=row_num, column=8).value = "Ναι" if animal.injured else "Όχι"
        ws.cell(row=row_num, column=9).value = animal.cage_number or ""
        ws.cell(row=row_num, column=10).value = animal.capture_location or ""
        ws.cell(row=row_num, column=11).value = animal.capture_date.strftime('%d/%m/%Y') if animal.capture_date else ""
        ws.cell(row=row_num, column=12).value = animal.entry_date.strftime('%d/%m/%Y') if animal.entry_date else ""
        ws.cell(row=row_num, column=13).value = animal.get_adoption_status_display()
        ws.cell(row=row_num, column=14).value = "Ναι" if animal.public_visibility else "Όχι"
        ws.cell(row=row_num, column=15).value = animal.shelter or ""
        ws.cell(row=row_num, column=16).value = str(animal.created_by) if animal.created_by else ""
        ws.cell(row=row_num, column=17).value = animal.created_at.strftime('%d/%m/%Y %H:%M') if animal.created_at else ""
    
    # Auto-size columns
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Enable filters
    ws.auto_filter.ref = ws.dimensions
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response


def export_vaccinations_to_excel(queryset, filename=None):
    """Export vaccinations queryset to Excel file"""
    
    if filename is None:
        filename = f'vaccinations_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Εμβολιασμοί"
    
    # Define headers
    headers = [
        'Όνομα Ζώου', 'Chip ID', 'Είδος',
        'Εμβόλιο', 'Σκεύασμα', 'Αρ. Παρτίδας',
        'Ημ/νία Χορήγησης', 'Επόμενη Ημ/νία', 'Κτηνίατρος',
        'Δημιουργήθηκε Από', 'Δημιουργήθηκε Στις'
    ]
    
    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_num, vaccination in enumerate(queryset.select_related('animal'), 2):
        ws.cell(row=row_num, column=1).value = vaccination.animal.name
        ws.cell(row=row_num, column=2).value = vaccination.animal.chip_id
        ws.cell(row=row_num, column=3).value = vaccination.animal.get_species_display()
        ws.cell(row=row_num, column=4).value = vaccination.get_vaccine_name_display()
        ws.cell(row=row_num, column=5).value = vaccination.other_vaccine_name or ""
        ws.cell(row=row_num, column=6).value = vaccination.batch_number or ""
        ws.cell(row=row_num, column=7).value = vaccination.date_administered.strftime('%d/%m/%Y') if vaccination.date_administered else ""
        ws.cell(row=row_num, column=8).value = vaccination.next_due_date.strftime('%d/%m/%Y') if vaccination.next_due_date else ""
        ws.cell(row=row_num, column=9).value = vaccination.administered_by or ""
        ws.cell(row=row_num, column=10).value = str(vaccination.created_by) if vaccination.created_by else ""
        ws.cell(row=row_num, column=11).value = vaccination.created_at.strftime('%d/%m/%Y %H:%M') if vaccination.created_at else ""
    
    # Auto-size columns
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Enable filters
    ws.auto_filter.ref = ws.dimensions
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response